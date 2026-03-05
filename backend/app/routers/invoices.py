from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from bson import ObjectId
from datetime import datetime, timedelta, timezone
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.platypus import KeepTogether

from ..database import db_helper
from ..models.invoice import InvoiceResponse, PaymentCreate, PaymentRecord, PaymentMethod, PaymentStatus
from ..models.user import TokenData
from ..utils.auth import require_admin
from ..utils.gst_invoice_excel import generate_gst_invoice_xlsx

router = APIRouter(prefix="/api/invoices", tags=["Invoices"])


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def _fix_doc(doc: dict) -> dict:
    """Convert ObjectId → str and ensure required defaults."""
    doc["_id"] = str(doc["_id"])
    doc.setdefault("items", [])
    doc.setdefault("payments", [])
    doc.setdefault("subtotal", 0)
    doc.setdefault("gst_amount", 0)
    doc.setdefault("total_amount", 0)
    doc.setdefault("amount_paid", 0)
    doc.setdefault("balance_due", 0)
    doc.setdefault("payment_status", PaymentStatus.PENDING.value)
    doc.setdefault("customer_name", "")
    doc.setdefault("invoice_number", "")
    return doc


def _to_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is not None:
            return parsed.astimezone(timezone.utc).replace(tzinfo=None)
        return parsed
    except Exception:
        return None


def _parse_filter_date(value: Optional[str], end_of_day: bool = False) -> Optional[datetime]:
    if not value:
        return None

    parsed = _to_datetime(value)
    if not parsed:
        return None

    parsed = parsed.replace(hour=0, minute=0, second=0, microsecond=0)
    if end_of_day:
        parsed = parsed + timedelta(days=1) - timedelta(microseconds=1)
    return parsed


async def _generate_ledger_data(
    db,
    customer_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    query = {}
    if customer_id:
        query["customer_id"] = customer_id

    start_dt = _parse_filter_date(from_date, end_of_day=False)
    end_dt = _parse_filter_date(to_date, end_of_day=True)

    raw_entries = []
    cursor = db.invoices.find(query).sort("created_at", 1)
    async for invoice in cursor:
        invoice_id = str(invoice.get("_id"))
        invoice_number = invoice.get("invoice_number", "")
        invoice_dt = _to_datetime(invoice.get("created_at")) or datetime.utcnow()
        total_amount = float(invoice.get("total_amount", 0) or 0)

        raw_entries.append({
            "date": invoice_dt,
            "particular": f"Invoice {invoice_number}",
            "debit": round(total_amount, 2),
            "credit": 0.0,
            "invoice_id": invoice_id,
            "invoice_number": invoice_number,
            "payment_method": None,
            "sort_key": 0,
        })

        for payment in invoice.get("payments", []) or []:
            payment_dt = _to_datetime(payment.get("payment_date")) or invoice_dt
            payment_amount = float(payment.get("amount", 0) or 0)
            method = payment.get("method") or "unknown"

            raw_entries.append({
                "date": payment_dt,
                "particular": f"Payment via {str(method).replace('_', ' ').title()}",
                "debit": 0.0,
                "credit": round(payment_amount, 2),
                "invoice_id": invoice_id,
                "invoice_number": invoice_number,
                "payment_method": method,
                "sort_key": 1,
            })

    raw_entries.sort(key=lambda x: (x["date"], x["sort_key"], x["invoice_number"]))

    opening_balance = 0.0
    if start_dt:
        for entry in raw_entries:
            if entry["date"] < start_dt:
                opening_balance += entry["debit"] - entry["credit"]

    filtered_entries = []
    for entry in raw_entries:
        if start_dt and entry["date"] < start_dt:
            continue
        if end_dt and entry["date"] > end_dt:
            continue
        filtered_entries.append(entry)

    running_balance = opening_balance
    ledger_rows = []
    for entry in filtered_entries:
        running_balance += entry["debit"] - entry["credit"]
        ledger_rows.append({
            "date": entry["date"].isoformat(),
            "particular": entry["particular"],
            "debit": round(entry["debit"], 2),
            "credit": round(entry["credit"], 2),
            "balance": round(running_balance, 2),
            "invoice_id": entry["invoice_id"],
            "invoice_number": entry["invoice_number"],
            "payment_method": entry["payment_method"],
        })

    total_debit = round(sum(row["debit"] for row in ledger_rows), 2)
    total_credit = round(sum(row["credit"] for row in ledger_rows), 2)
    closing_balance = round(opening_balance + total_debit - total_credit, 2)

    customer_name = "All Customers"
    if customer_id:
        user_doc = None
        try:
            user_doc = await db.users.find_one({"_id": ObjectId(customer_id)})
        except Exception:
            user_doc = await db.users.find_one({"_id": customer_id})

        if user_doc:
            customer_name = user_doc.get("full_name") or user_doc.get("company_name") or customer_id
        else:
            customer_name = customer_id

    return {
        "customer_id": customer_id,
        "customer_name": customer_name,
        "from_date": from_date,
        "to_date": to_date,
        "opening_balance": round(opening_balance, 2),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "closing_balance": closing_balance,
        "entries": ledger_rows,
    }


def _build_ledger_pdf(ledger_data: dict) -> BytesIO:
    def _fmt_amount(value: float, with_suffix: bool = False) -> str:
        amount = float(value or 0)
        base = f"{abs(amount):,.2f}"
        if not with_suffix:
            return base
        return f"{base} {'Dr' if amount >= 0 else 'Cr'}"

    def _fmt_date_short(value) -> str:
        parsed = _to_datetime(value)
        if not parsed:
            return "-"
        return parsed.strftime("%d/%m/%y")

    entries = ledger_data.get("entries", []) or []
    entry_dates = [_to_datetime(e.get("date")) for e in entries if _to_datetime(e.get("date"))]

    from_raw = ledger_data.get("from_date")
    to_raw = ledger_data.get("to_date")
    from_dt = _to_datetime(from_raw) if from_raw else (min(entry_dates) if entry_dates else datetime.utcnow())
    to_dt = _to_datetime(to_raw) if to_raw else (max(entry_dates) if entry_dates else datetime.utcnow())

    from_text = from_dt.strftime("%d/%m/%Y") if from_dt else "-"
    to_text = to_dt.strftime("%d/%m/%Y") if to_dt else "-"

    styles = getSampleStyleSheet()
    header_title = ParagraphStyle("LedgerHeaderTitle", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER)
    header_sub = ParagraphStyle("LedgerHeaderSub", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8.5, alignment=TA_CENTER)
    header_state = ParagraphStyle("LedgerHeaderState", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, alignment=TA_CENTER)
    statement_title = ParagraphStyle("LedgerStatementTitle", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9, alignment=TA_CENTER)
    period_style = ParagraphStyle("LedgerPeriod", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.8, alignment=TA_LEFT)
    customer_style = ParagraphStyle("LedgerCustomer", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9, alignment=TA_CENTER)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    story = []

    story.append(Paragraph("R. R. ENTERPRISES", header_title))
    story.append(Paragraph("SHOP NO. 83/2, VILL SHAHIMAJRA", header_title))
    story.append(Paragraph("DISTT MOHALI", header_title))
    story.append(Paragraph("Punjab", header_state))
    story.append(Paragraph("Ledger Statement", statement_title))

    period_table = Table(
        [[
            Paragraph(f"FROM : {from_text} TO : {to_text}", period_style),
            Paragraph("Page: 1", ParagraphStyle("LedgerPage", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.8, alignment=TA_RIGHT)),
        ]],
        colWidths=[128 * mm, 52 * mm],
    )
    period_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(period_table)
    story.append(Spacer(1, 4))

    table_data = [
        ["Date", "Particulars", "Debit", "Credit", "Balance"],
        ["", Paragraph(ledger_data.get("customer_name", "All Customers").upper(), customer_style), "", "", _fmt_amount(ledger_data.get("opening_balance", 0), with_suffix=True)],
        ["", "Opening Balance", "", "", _fmt_amount(ledger_data.get("opening_balance", 0), with_suffix=True)],
    ]

    if entries:
        for row in entries:
            table_data.append([
                _fmt_date_short(row.get("date")),
                row.get("particular", ""),
                _fmt_amount(row.get("debit", 0)) if row.get("debit", 0) else "",
                _fmt_amount(row.get("credit", 0)) if row.get("credit", 0) else "",
                _fmt_amount(row.get("balance", 0), with_suffix=True),
            ])
    else:
        table_data.append(["", "No transactions in selected period", "", "", _fmt_amount(ledger_data.get("closing_balance", 0), with_suffix=True)])

    table_data.append(["", "", "", "", ""])
    table_data.append([
        "",
        "GRAND TOTAL",
        _fmt_amount(ledger_data.get("total_debit", 0)),
        _fmt_amount(ledger_data.get("total_credit", 0)),
        _fmt_amount(ledger_data.get("closing_balance", 0), with_suffix=True),
    ])

    col_widths = [22 * mm, 86 * mm, 24 * mm, 24 * mm, 28 * mm]
    ledger_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    last_row = len(table_data) - 1
    ledger_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTNAME", (1, last_row), (-1, last_row), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (1, -1), "LEFT"),
        ("ALIGN", (2, 0), (4, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.HexColor("#444444")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#444444")),
        ("LINEBELOW", (0, 1), (-1, 1), 0.6, colors.HexColor("#666666")),
        ("LINEABOVE", (0, last_row), (-1, last_row), 0.8, colors.HexColor("#444444")),
        ("LINEBELOW", (0, last_row), (-1, last_row), 1.0, colors.HexColor("#444444")),

        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    story.append(ledger_table)
    doc.build(story)
    buf.seek(0)
    return buf


async def _enrich_items_with_docket(db, doc: dict) -> dict:
    """Fill missing docket numbers on invoice items from consignments/shipments."""
    items = doc.get("items", []) or []
    updated = False

    for item in items:
        if item.get("docket_no"):
            continue

        docket_no = None
        shipment_id = item.get("shipment_id")
        tracking_number = item.get("tracking_number")

        if shipment_id:
            consignment = await db.consignments.find_one({"shipment_id": shipment_id})
            if consignment:
                docket_no = consignment.get("docket_no") or consignment.get("consignment_no")
        if not docket_no and tracking_number:
            shipment = await db.shipments.find_one({"tracking_number": tracking_number})
            if shipment:
                docket_no = shipment.get("docket_no")

        if docket_no:
            item["docket_no"] = docket_no
            updated = True

    if updated:
        await db.invoices.update_one(
            {"_id": ObjectId(doc["_id"])},
            {"$set": {"items": items, "updated_at": datetime.utcnow()}}
        )
        doc["items"] = items

    return doc


# ─────────────────────────────────────────────
#  CRUD
# ─────────────────────────────────────────────

@router.get("/", response_model=List[InvoiceResponse])
async def list_invoices(
    customer_id: Optional[str] = None,
    payment_status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    token_data: TokenData = Depends(require_admin),
):
    """List all invoices with optional filters (Admin only)."""
    db = db_helper.db
    query: dict = {}
    if customer_id:
        query["customer_id"] = customer_id
    if payment_status:
        query["payment_status"] = payment_status

    cursor = db.invoices.find(query).sort("created_at", -1).skip(skip).limit(limit)
    result = []
    async for doc in cursor:
        try:
            fixed = _fix_doc(doc)
            fixed = await _enrich_items_with_docket(db, fixed)
            result.append(InvoiceResponse(**fixed))
        except Exception as e:
            print(f"Skipping invoice {doc.get('_id')}: {e}")
    return result


@router.get("/ledger")
async def get_customer_ledger(
    customer_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    token_data: TokenData = Depends(require_admin),
):
    """Get customer ledger statement built from invoice debits and payment credits."""
    db = db_helper.db
    try:
        return await _generate_ledger_data(
            db=db,
            customer_id=customer_id,
            from_date=from_date,
            to_date=to_date,
        )
    except Exception as exc:
        print(f"Ledger API failed: {exc}")
        raise HTTPException(status_code=500, detail="Failed to generate ledger statement")


@router.get("/ledger/excel")
async def export_customer_ledger_excel(
    customer_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    token_data: TokenData = Depends(require_admin),
):
    """Export customer ledger statement to Excel."""
    db = db_helper.db
    try:
        ledger_data = await _generate_ledger_data(
            db=db,
            customer_id=customer_id,
            from_date=from_date,
            to_date=to_date,
        )
    except Exception as exc:
        print(f"Ledger Excel export failed: {exc}")
        raise HTTPException(status_code=500, detail="Failed to export ledger excel")

    def _fmt_amount(value: float) -> str:
        return f"{float(value or 0):,.2f}"

    def _fmt_balance(value: float) -> str:
        amount = float(value or 0)
        return f"{abs(amount):,.2f} {'Dr' if amount >= 0 else 'Cr'}"

    entries = ledger_data.get("entries", []) or []
    entry_dates = [_to_datetime(e.get("date")) for e in entries if _to_datetime(e.get("date"))]
    from_dt = _to_datetime(ledger_data.get("from_date")) if ledger_data.get("from_date") else (min(entry_dates) if entry_dates else datetime.utcnow())
    to_dt = _to_datetime(ledger_data.get("to_date")) if ledger_data.get("to_date") else (max(entry_dates) if entry_dates else datetime.utcnow())
    from_text = from_dt.strftime("%d/%m/%Y") if from_dt else "-"
    to_text = to_dt.strftime("%d/%m/%Y") if to_dt else "-"

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger Statement"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    bold_center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="666666")
    medium = Side(style="medium", color="444444")

    def _merge_title(row_no: int, text: str, size: int = 12):
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=5)
        cell = ws.cell(row=row_no, column=1, value=text)
        cell.font = Font(name="Calibri", bold=True, size=size)
        cell.alignment = bold_center

    _merge_title(1, "R. R. ENTERPRISES", 14)
    _merge_title(2, "SHOP NO. 83/2, VILL SHAHIMAJRA", 12)
    _merge_title(3, "DISTT MOHALI", 12)
    _merge_title(4, "Punjab", 12)
    _merge_title(5, "Ledger Statement", 12)

    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=4)
    ws.cell(row=6, column=1, value=f"FROM : {from_text} TO : {to_text}")
    ws.cell(row=6, column=1).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=6, column=1).alignment = left_align
    ws.cell(row=6, column=5, value="Page: 1")
    ws.cell(row=6, column=5).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=6, column=5).alignment = right_align

    header_row = 8
    headers = ["Date", "Particulars", "Debit", "Credit", "Balance"]
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=title)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = left_align if idx <= 2 else right_align
        cell.border = Border(top=medium, bottom=medium)

    customer_row = 9
    ws.merge_cells(start_row=customer_row, start_column=2, end_row=customer_row, end_column=4)
    customer_cell = ws.cell(row=customer_row, column=2, value=(ledger_data.get("customer_name", "All Customers") or "All Customers").upper())
    customer_cell.font = Font(name="Calibri", bold=True, size=11)
    customer_cell.alignment = bold_center
    ws.cell(row=customer_row, column=5, value=_fmt_balance(ledger_data.get("opening_balance", 0)))
    ws.cell(row=customer_row, column=5).alignment = right_align
    ws.cell(row=customer_row, column=5).font = Font(name="Calibri", bold=True, size=10)

    opening_row = 10
    ws.cell(row=opening_row, column=2, value="Opening Balance")
    ws.cell(row=opening_row, column=2).alignment = left_align
    ws.cell(row=opening_row, column=5, value=_fmt_balance(ledger_data.get("opening_balance", 0)))
    ws.cell(row=opening_row, column=5).alignment = right_align

    current_row = 11
    if entries:
        for item in entries:
            date_obj = _to_datetime(item.get("date"))
            ws.cell(row=current_row, column=1, value=date_obj.strftime("%d/%m/%y") if date_obj else "-")
            ws.cell(row=current_row, column=2, value=item.get("particular", ""))
            ws.cell(row=current_row, column=3, value=_fmt_amount(item.get("debit", 0)) if item.get("debit", 0) else "")
            ws.cell(row=current_row, column=4, value=_fmt_amount(item.get("credit", 0)) if item.get("credit", 0) else "")
            ws.cell(row=current_row, column=5, value=_fmt_balance(item.get("balance", 0)))

            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=2).alignment = left_align
            ws.cell(row=current_row, column=3).alignment = right_align
            ws.cell(row=current_row, column=4).alignment = right_align
            ws.cell(row=current_row, column=5).alignment = right_align
            current_row += 1
    else:
        ws.cell(row=current_row, column=2, value="No transactions in selected period")
        ws.cell(row=current_row, column=5, value=_fmt_balance(ledger_data.get("closing_balance", 0)))
        ws.cell(row=current_row, column=2).alignment = left_align
        ws.cell(row=current_row, column=5).alignment = right_align
        current_row += 1

    current_row += 1
    ws.cell(row=current_row, column=2, value="GRAND TOTAL")
    ws.cell(row=current_row, column=3, value=_fmt_amount(ledger_data.get("total_debit", 0)))
    ws.cell(row=current_row, column=4, value=_fmt_amount(ledger_data.get("total_credit", 0)))
    ws.cell(row=current_row, column=5, value=_fmt_balance(ledger_data.get("closing_balance", 0)))

    for col in range(1, 6):
        c = ws.cell(row=current_row, column=col)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.border = Border(top=medium, bottom=medium)
        c.alignment = left_align if col <= 2 else right_align

    for row_no in range(9, current_row + 1):
        for col_no in range(1, 6):
            if row_no == current_row:
                continue
            cell = ws.cell(row=row_no, column=col_no)
            if cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style:
                continue
            cell.border = Border(bottom=thin) if row_no in (9,) else Border()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_customer = (ledger_data.get("customer_name") or "all_customers").replace(" ", "_").lower()
    filename = f"ledger_statement_{file_customer}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ledger/pdf")
async def export_customer_ledger_pdf(
    customer_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    token_data: TokenData = Depends(require_admin),
):
    """Export customer ledger statement to PDF."""
    db = db_helper.db
    try:
        ledger_data = await _generate_ledger_data(
            db=db,
            customer_id=customer_id,
            from_date=from_date,
            to_date=to_date,
        )
        pdf_buf = _build_ledger_pdf(ledger_data)
    except Exception as exc:
        print(f"Ledger PDF export failed: {exc}")
        raise HTTPException(status_code=500, detail="Failed to export ledger pdf")

    file_customer = (ledger_data.get("customer_name") or "all_customers").replace(" ", "_").lower()
    filename = f"ledger_statement_{file_customer}.pdf"
    return StreamingResponse(
        pdf_buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(
    invoice_id: str,
    token_data: TokenData = Depends(require_admin),
):
    """Get a single invoice by ID."""
    db = db_helper.db
    doc = await db.invoices.find_one({"_id": ObjectId(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Invoice not found")
    fixed = _fix_doc(doc)
    fixed = await _enrich_items_with_docket(db, fixed)
    return InvoiceResponse(**fixed)


@router.post("/{invoice_id}/payment")
async def add_payment(
    invoice_id: str,
    payment: PaymentCreate,
    token_data: TokenData = Depends(require_admin),
):
    """Record a payment against an invoice."""
    db = db_helper.db
    doc = await db.invoices.find_one({"_id": ObjectId(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Invoice not found")

    balance_due = doc.get("balance_due", 0)
    if payment.amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be positive")
    if payment.amount > balance_due + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Payment ₹{payment.amount:.2f} exceeds balance due ₹{balance_due:.2f}",
        )

    payment_record = {
        "amount": payment.amount,
        "method": payment.method.value if hasattr(payment.method, "value") else payment.method,
        "transaction_ref": payment.transaction_ref or "",
        "payment_date": datetime.utcnow(),
        "received_by": token_data.user_id,
        "notes": payment.notes or "",
    }

    new_paid = doc.get("amount_paid", 0) + payment.amount
    new_balance = doc.get("total_amount", 0) - new_paid
    new_status = (
        PaymentStatus.PAID.value
        if new_balance <= 0.01
        else PaymentStatus.PARTIAL.value
    )

    await db.invoices.update_one(
        {"_id": ObjectId(invoice_id)},
        {
            "$push": {"payments": payment_record},
            "$set": {
                "amount_paid": round(new_paid, 2),
                "balance_due": round(max(new_balance, 0), 2),
                "payment_status": new_status,
                "updated_at": datetime.utcnow(),
            },
        },
    )
    return {"message": "Payment recorded", "new_status": new_status}


# ─────────────────────────────────────────────
#  PDF Download  (reportlab)
# ─────────────────────────────────────────────

def _build_pdf(invoice: dict) -> BytesIO:
    """Generate A4 GST invoice PDF matching the Excel grid format."""
    from decimal import Decimal
    from ..utils.gst_invoice_excel import amount_to_words_indian

    def _fmt_date(value) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        text = str(value or "").strip()
        if "T" in text:
            try:
                return datetime.fromisoformat(text.split(".")[0]).strftime("%d/%m/%Y")
            except Exception:
                return text
        return text

    def _f(value, decimals: int = 2) -> float:
        try:
            return round(float(value or 0), decimals)
        except Exception:
            return 0.0

    def _build_invoice_data(doc: dict) -> dict:
        created_at = doc.get("created_at")
        invoice_date = created_at if isinstance(created_at, datetime) else datetime.utcnow()

        subtotal = _f(doc.get("subtotal", 0))
        gst_amount = _f(doc.get("gst_amount", 0))
        effective_gst_rate = round((gst_amount / subtotal) * 100, 2) if subtotal > 0 else 18.0

        line_items = []
        for item in doc.get("items", []) or []:
            description = item.get("description", "") or "Service"
            amount = _f(item.get("amount", 0))
            line_items.append(
                {
                    "name": description,
                    "uom": "Nos.",
                    "ac_code": "9968",
                    "rate": amount,
                    "qty": 1,
                    "gst_rate": effective_gst_rate,
                }
            )

        return {
            "invoice_no": doc.get("invoice_number", ""),
            "invoice_date": invoice_date,
            "seller_details": {
                "company_name": "R. R. ENTERPRISES",
                "address_lines": [
                    "ADDRESS-C-102/A3) PHASE -7 INDL AREA MOHALI",
                    "Mob. No. : 9888483817-9592410333",
                    "GSTIN : 03CDXPD2324L1ZH",
                    "E-Mail : RRENTERPRISES922@HOTMAIL.COM",
                ],
                "state": "PUNJAB",
                "state_code": "03",
                "gstin": "03CDXPD2324L1ZH",
                "email": "RRENTERPRISES922@HOTMAIL.COM",
            },
            "buyer_details": {
                "name": doc.get("customer_name", ""),
                "address": doc.get("billing_address", ""),
                "gstin": doc.get("customer_gstin", ""),
                "state": doc.get("customer_state") or doc.get("state") or "PUNJAB",
                "state_code": str(doc.get("customer_state_code") or doc.get("state_code") or "03"),
            },
            "items": line_items,
            "bank_details": {
                "bank_name": "INDIAN BANK",
                "branch": "PHASE-1 MOHALI",
                "ifsc_code": "IDIB000P637",
                "account_no": "50267738489",
                "account_type": "CURRENT A/C",
            },
        }

    def _compute_lines(data: dict):
        seller_code = str(data.get("seller_details", {}).get("state_code", ""))
        buyer_code = str(data.get("buyer_details", {}).get("state_code", ""))
        intrastate = bool(seller_code and buyer_code and seller_code == buyer_code)

        rows = []
        totals = {
            "taxable": 0.0,
            "cgst": 0.0,
            "sgst": 0.0,
            "igst": 0.0,
            "grand": 0.0,
        }

        for idx, item in enumerate(data.get("items", []) or [], start=1):
            qty = _f(item.get("qty", 1), 4)
            rate = _f(item.get("rate", 0))
            gst_rate = _f(item.get("gst_rate", 18))
            taxable = _f(qty * rate)

            if intrastate:
                cgst_rate = _f(gst_rate / 2)
                sgst_rate = _f(gst_rate / 2)
                igst_rate = 0.0
            else:
                cgst_rate = 0.0
                sgst_rate = 0.0
                igst_rate = gst_rate

            cgst_amt = _f(taxable * cgst_rate / 100)
            sgst_amt = _f(taxable * sgst_rate / 100)
            igst_amt = _f(taxable * igst_rate / 100)
            total = _f(taxable + cgst_amt + sgst_amt + igst_amt)

            totals["taxable"] += taxable
            totals["cgst"] += cgst_amt
            totals["sgst"] += sgst_amt
            totals["igst"] += igst_amt
            totals["grand"] += total

            rows.append(
                {
                    "sr": idx,
                    "name": item.get("name", ""),
                    "uom": item.get("uom", "Nos."),
                    "ac_code": item.get("ac_code", ""),
                    "rate": rate,
                    "taxable": taxable,
                    "cgst_rate": cgst_rate,
                    "cgst_amt": cgst_amt,
                    "sgst_rate": sgst_rate,
                    "sgst_amt": sgst_amt,
                    "igst_rate": igst_rate,
                    "igst_amt": igst_amt,
                    "total": total,
                }
            )

        for k in totals:
            totals[k] = _f(totals[k])

        return rows, totals

    invoice_data = _build_invoice_data(invoice)
    item_rows, totals = _compute_lines(invoice_data)

    seller = invoice_data["seller_details"]
    buyer = invoice_data["buyer_details"]
    bank = invoice_data["bank_details"]

    header_lines = list(seller.get("address_lines") or [])
    while len(header_lines) < 4:
        header_lines.append("")

    def blank_row():
        return [""] * 13

    body_style = ParagraphStyle("gst_body", fontName="Helvetica", fontSize=7.6, leading=9.2, wordWrap="CJK")
    body_bold = ParagraphStyle("gst_bold", fontName="Helvetica-Bold", fontSize=7.8, leading=9.2, wordWrap="CJK")
    label_bold = ParagraphStyle("gst_label", fontName="Helvetica-Bold", fontSize=7.8, leading=9.2)
    head_bold = ParagraphStyle("gst_head", fontName="Helvetica-Bold", fontSize=9, leading=10, alignment=TA_CENTER)

    def _p(value, style=body_style):
        text = str(value or "").replace("\n", "<br/>")
        return Paragraph(text, style)

    data = []

    r = blank_row()
    r[0] = Paragraph(str(seller.get("company_name", "R. R. ENTERPRISES")).upper(), ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=14, alignment=TA_CENTER))
    data.append(r)

    for line in header_lines[:4]:
        rr = blank_row()
        rr[0] = Paragraph(str(line), body_bold)
        data.append(rr)

    r = blank_row()
    r[0] = _p("Invoice No.", label_bold)
    r[1] = _p(invoice_data.get("invoice_no", ""))
    r[6] = _p("Details of Receiver", head_bold)
    data.append(r)

    r = blank_row()
    r[0] = _p("Invoice Date", label_bold)
    r[1] = _p(_fmt_date(invoice_data.get("invoice_date")))
    r[6] = _p("Name", label_bold)
    r[7] = _p(buyer.get("name", ""))
    data.append(r)

    r = blank_row()
    r[0] = _p("State", label_bold)
    r[1] = _p(seller.get("state", ""))
    r[3] = _p("State Code", label_bold)
    r[4] = _p(seller.get("state_code", ""))
    r[6] = _p("Address", label_bold)
    r[7] = _p(buyer.get("address", ""))
    data.append(r)

    r = blank_row()
    r[6] = _p("GSTIN", label_bold)
    r[7] = _p(buyer.get("gstin", ""))
    data.append(r)

    r = blank_row()
    r[6] = _p("State", label_bold)
    r[7] = _p(buyer.get("state", ""))
    r[9] = _p("State Code", label_bold)
    r[10] = _p(buyer.get("state_code", ""))
    data.append(r)

    r = blank_row()
    r[0] = _p("Sr. No", body_bold)
    r[1] = _p("Name of Product/Service", body_bold)
    r[2] = _p("UOM", body_bold)
    r[3] = _p("AC CODE", body_bold)
    r[4] = _p("Rate", body_bold)
    r[5] = _p("Taxable Value", body_bold)
    r[6] = _p("CGST", head_bold)
    r[8] = _p("SGST", head_bold)
    r[10] = _p("IGST", head_bold)
    r[12] = _p("Total (₹)", body_bold)
    data.append(r)

    r = blank_row()
    r[6] = _p("Rate", body_bold)
    r[7] = _p("Amount", body_bold)
    r[8] = _p("Rate", body_bold)
    r[9] = _p("Amount", body_bold)
    r[10] = _p("Rate", body_bold)
    r[11] = _p("Amount", body_bold)
    data.append(r)

    if not item_rows:
        item_rows = [
            {
                "sr": 1,
                "name": "",
                "uom": "",
                "ac_code": "",
                "rate": 0,
                "taxable": 0,
                "cgst_rate": 0,
                "cgst_amt": 0,
                "sgst_rate": 0,
                "sgst_amt": 0,
                "igst_rate": 0,
                "igst_amt": 0,
                "total": 0,
            }
        ]

    for item in item_rows:
        row = [
            _p(item["sr"]),
            _p(item["name"]),
            _p(item["uom"]),
            _p(item["ac_code"]),
            _p(f"{_f(item['rate']):,.2f}"),
            _p(f"{_f(item['taxable']):,.2f}"),
            _p(f"{_f(item['cgst_rate']):.2f}%"),
            _p(f"{_f(item['cgst_amt']):,.2f}"),
            _p(f"{_f(item['sgst_rate']):.2f}%"),
            _p(f"{_f(item['sgst_amt']):,.2f}"),
            _p(f"{_f(item['igst_rate']):.2f}%"),
            _p(f"{_f(item['igst_amt']):,.2f}"),
            _p(f"{_f(item['total']):,.2f}"),
        ]
        data.append(row)

    total_row_idx = len(data)
    r = blank_row()
    r[0] = _p("Total", body_bold)
    r[5] = _p(f"{_f(totals['taxable']):,.2f}", body_bold)
    r[7] = _p(f"{_f(totals['cgst']):,.2f}", body_bold)
    r[9] = _p(f"{_f(totals['sgst']):,.2f}", body_bold)
    r[11] = _p(f"{_f(totals['igst']):,.2f}", body_bold)
    r[12] = _p(f"{_f(totals['grand']):,.2f}", body_bold)
    data.append(r)

    words = amount_to_words_indian(Decimal(str(_f(totals["grand"]))))
    r = blank_row()
    r[0] = _p(f"Total Invoice Amount in Words : {words}", body_bold)
    r[7] = _p("Total Amount before Tax", body_bold)
    r[12] = _p(f"{_f(totals['taxable']):,.2f}", body_bold)
    data.append(r)

    bottom_rows = [
        (f"Bank Name : {bank.get('bank_name', '')}", "Add: CGST", totals["cgst"]),
        (f"Branch : {bank.get('branch', '')}", "Add: SGST", totals["sgst"]),
        (f"IFSC Code : {bank.get('ifsc_code', '')}", "Add: IGST", totals["igst"]),
        (f"A/C No : {bank.get('account_no', '')}", "Total Amount after Tax", totals["grand"]),
        (f"Type of A/C : {bank.get('account_type', '')}", "", ""),
    ]

    for left_text, right_label, right_val in bottom_rows:
        rr = blank_row()
        rr[0] = _p(left_text)
        rr[7] = _p(right_label, body_bold if right_label in ("Total Amount before Tax", "Total Amount after Tax") else body_style)
        rr[12] = _p(
            f"{_f(right_val):,.2f}" if right_val != "" else "",
            body_bold if right_label == "Total Amount after Tax" else body_style,
        )
        data.append(rr)

    rr = blank_row()
    rr[0] = _p(
        "Terms & Conditions: Goods once sold will not be taken back. Please settle dues within 15 days.",
        body_style,
    )
    data.append(rr)

    rr = blank_row()
    rr[7] = _p("For R. R. ENTERPRISES", body_bold)
    data.append(rr)

    data.append(blank_row())

    rr = blank_row()
    rr[7] = _p("Authorised Signatory", body_bold)
    data.append(rr)

    doc = SimpleDocTemplate(
        BytesIO(),
        pagesize=A4,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    usable_w = A4[0] - doc.leftMargin - doc.rightMargin

    weights = [9, 27, 8, 10, 11, 13, 8, 13, 8, 13, 8, 13, 14]
    total_weight = sum(weights)
    col_widths = [usable_w * (w / total_weight) for w in weights]

    table = Table(data, colWidths=col_widths, repeatRows=2)

    style = [
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),

        ("SPAN", (0, 0), (12, 0)),
        ("SPAN", (0, 1), (12, 1)),
        ("SPAN", (0, 2), (12, 2)),
        ("SPAN", (0, 3), (12, 3)),
        ("SPAN", (0, 4), (12, 4)),

        ("SPAN", (1, 5), (5, 5)),
        ("SPAN", (1, 6), (5, 6)),
        ("SPAN", (1, 7), (2, 7)),
        ("SPAN", (4, 7), (5, 7)),
        ("SPAN", (6, 5), (12, 5)),
        ("SPAN", (7, 6), (12, 6)),
        ("SPAN", (7, 7), (12, 7)),
        ("SPAN", (7, 8), (12, 8)),
        ("SPAN", (7, 9), (8, 9)),
        ("SPAN", (10, 9), (12, 9)),

        ("SPAN", (0, 10), (0, 11)),
        ("SPAN", (1, 10), (1, 11)),
        ("SPAN", (2, 10), (2, 11)),
        ("SPAN", (3, 10), (3, 11)),
        ("SPAN", (4, 10), (4, 11)),
        ("SPAN", (5, 10), (5, 11)),
        ("SPAN", (6, 10), (7, 10)),
        ("SPAN", (8, 10), (9, 10)),
        ("SPAN", (10, 10), (11, 10)),
        ("SPAN", (12, 10), (12, 11)),

        ("SPAN", (0, total_row_idx), (4, total_row_idx)),
        ("SPAN", (0, total_row_idx + 1), (6, total_row_idx + 1)),
        ("SPAN", (7, total_row_idx + 1), (11, total_row_idx + 1)),
        ("SPAN", (0, total_row_idx + 2), (6, total_row_idx + 2)),
        ("SPAN", (7, total_row_idx + 2), (11, total_row_idx + 2)),
        ("SPAN", (0, total_row_idx + 3), (6, total_row_idx + 3)),
        ("SPAN", (7, total_row_idx + 3), (11, total_row_idx + 3)),
        ("SPAN", (0, total_row_idx + 4), (6, total_row_idx + 4)),
        ("SPAN", (7, total_row_idx + 4), (11, total_row_idx + 4)),
        ("SPAN", (0, total_row_idx + 5), (6, total_row_idx + 5)),
        ("SPAN", (7, total_row_idx + 5), (11, total_row_idx + 5)),
        ("SPAN", (0, total_row_idx + 6), (6, total_row_idx + 6)),
        ("SPAN", (0, total_row_idx + 7), (6, total_row_idx + 7)),
        ("SPAN", (7, total_row_idx + 7), (12, total_row_idx + 7)),
        ("SPAN", (0, total_row_idx + 8), (6, total_row_idx + 8)),
        ("SPAN", (7, total_row_idx + 8), (12, total_row_idx + 8)),
        ("SPAN", (0, total_row_idx + 9), (6, total_row_idx + 9)),
        ("SPAN", (7, total_row_idx + 9), (12, total_row_idx + 9)),

        ("ALIGN", (0, 0), (12, 4), "CENTER"),
        ("ALIGN", (0, 10), (12, 11), "CENTER"),
        ("ALIGN", (0, 12), (3, total_row_idx), "CENTER"),
        ("ALIGN", (4, 12), (12, total_row_idx + 5), "RIGHT"),
        ("ALIGN", (0, total_row_idx), (0, total_row_idx), "RIGHT"),
        ("ALIGN", (0, total_row_idx + 1), (6, total_row_idx + 9), "LEFT"),
        ("ALIGN", (7, total_row_idx + 1), (12, total_row_idx + 9), "RIGHT"),
        ("ALIGN", (7, total_row_idx + 7), (12, total_row_idx + 9), "RIGHT"),

        ("FONTNAME", (0, 0), (12, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 5), (12, 11), "Helvetica-Bold"),
        ("FONTNAME", (0, total_row_idx), (12, total_row_idx), "Helvetica-Bold"),
    ]

    table.setStyle(TableStyle(style))

    out = BytesIO()
    final_doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    final_doc.build([table])
    out.seek(0)
    return out


@router.get("/{invoice_id}/pdf")
async def download_invoice_pdf(
    invoice_id: str,
    token_data: TokenData = Depends(require_admin),
):
    """Download invoice as a professionally formatted PDF."""
    db = db_helper.db
    doc = await db.invoices.find_one({"_id": ObjectId(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Invoice not found")

    doc = _fix_doc(doc)
    doc = await _enrich_items_with_docket(db, doc)
    buf = _build_pdf(doc)
    filename = f"invoice_{doc.get('invoice_number', invoice_id)}.pdf"

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
#  Excel Download
# ─────────────────────────────────────────────

@router.get("/{invoice_id}/excel")
async def download_invoice_excel(
    invoice_id: str,
    token_data: TokenData = Depends(require_admin),
):
    """Download invoice as formatted GST Excel (A4 print-ready)."""
    db = db_helper.db
    doc = await db.invoices.find_one({"_id": ObjectId(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Invoice not found")

    doc = _fix_doc(doc)
    doc = await _enrich_items_with_docket(db, doc)

    created_at = doc.get("created_at")
    invoice_date = created_at if isinstance(created_at, datetime) else datetime.utcnow()

    subtotal = float(doc.get("subtotal", 0) or 0)
    gst_amount = float(doc.get("gst_amount", 0) or 0)
    effective_gst_rate = round((gst_amount / subtotal) * 100, 2) if subtotal > 0 else 18.0

    seller_details = {
        "company_name": "R. R. ENTERPRISES",
        "address_lines": [
            "ADDRESS-C-102/A3) PHASE -7 INDL AREA MOHALI",
            "Mob. No. : 9888483817-9592410333",
            "GSTIN : 03CDXPD2324L1ZH",
            "E-Mail : RRENTERPRISES922@HOTMAIL.COM",
        ],
        "state": "PUNJAB",
        "state_code": "03",
        "gstin": "03CDXPD2324L1ZH",
        "email": "RRENTERPRISES922@HOTMAIL.COM",
    }

    buyer_state = doc.get("customer_state") or doc.get("state") or "PUNJAB"
    buyer_state_code = str(doc.get("customer_state_code") or doc.get("state_code") or "03")

    buyer_details = {
        "name": doc.get("customer_name", ""),
        "address": doc.get("billing_address", ""),
        "gstin": doc.get("customer_gstin", ""),
        "state": buyer_state,
        "state_code": buyer_state_code,
    }

    line_items = []
    for item in doc.get("items", []) or []:
        description = item.get("description", "") or "Service"
        amount = float(item.get("amount", 0) or 0)
        line_items.append(
            {
                "name": description,
                "uom": "Nos.",
                "ac_code": "9968",
                "rate": amount,
                "qty": 1,
                "gst_rate": effective_gst_rate,
            }
        )

    invoice_data = {
        "invoice_no": doc.get("invoice_number", invoice_id),
        "invoice_date": invoice_date,
        "seller_details": seller_details,
        "buyer_details": buyer_details,
        "items": line_items,
        "bank_details": {
            "bank_name": "INDIAN BANK",
            "branch": "PHASE-1 MOHALI",
            "ifsc_code": "IDIB000P637",
            "account_no": "50267738489",
            "account_type": "CURRENT A/C",
        },
    }

    from tempfile import TemporaryDirectory
    from pathlib import Path

    with TemporaryDirectory() as tmp_dir:
        output_path = generate_gst_invoice_xlsx(invoice_data, output_dir=tmp_dir)
        buf = BytesIO(Path(output_path).read_bytes())
        buf.seek(0)

    filename = f"invoice_{doc.get('invoice_number', invoice_id)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
