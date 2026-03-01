from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins


THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def _d(value: Any) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _safe_state_code(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.isdigit() and len(text) == 1:
        return f"0{text}"
    return text


def _format_invoice_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value or "")


ONES = [
    "",
    "one",
    "two",
    "three",
    "four",
    "five",
    "six",
    "seven",
    "eight",
    "nine",
    "ten",
    "eleven",
    "twelve",
    "thirteen",
    "fourteen",
    "fifteen",
    "sixteen",
    "seventeen",
    "eighteen",
    "nineteen",
]

TENS = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]


def _num_to_words_1_to_999(number: int) -> str:
    if number == 0:
        return ""
    parts: list[str] = []
    if number >= 100:
        parts.append(f"{ONES[number // 100]} hundred")
        number %= 100
    if 0 < number < 20:
        parts.append(ONES[number])
    elif number >= 20:
        parts.append(TENS[number // 10])
        if number % 10:
            parts.append(ONES[number % 10])
    return " ".join([part for part in parts if part]).strip()


def amount_to_words_indian(amount: Decimal) -> str:
    amount = _d(amount)
    rupees = int(amount)
    paise = int((amount - Decimal(rupees)) * 100)

    if rupees == 0:
        rupee_words = "zero"
    else:
        crore = rupees // 10000000
        rupees %= 10000000
        lakh = rupees // 100000
        rupees %= 100000
        thousand = rupees // 1000
        rupees %= 1000
        rest = rupees

        chunks: list[str] = []
        if crore:
            chunks.append(f"{_num_to_words_1_to_999(crore)} crore")
        if lakh:
            chunks.append(f"{_num_to_words_1_to_999(lakh)} lakh")
        if thousand:
            chunks.append(f"{_num_to_words_1_to_999(thousand)} thousand")
        if rest:
            chunks.append(_num_to_words_1_to_999(rest))
        rupee_words = " ".join(chunks).strip()

    if paise:
        paise_words = _num_to_words_1_to_999(paise)
        return f"{rupee_words} rupees and {paise_words} paise only".title()
    return f"{rupee_words} rupees only".title()


def _set_cell(
    ws,
    row: int,
    col: int,
    value: Any = "",
    *,
    font: Font | None = None,
    align: Alignment | None = None,
    border: Border | None = THIN_BORDER,
):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _set_border_range(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def _merge_and_set(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    value: Any,
    *,
    font: Font | None = None,
    align: Alignment | None = None,
):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    _set_cell(ws, start_row, start_col, value=value, font=font, align=align)


def _prepare_items(invoice_data: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Decimal], bool]:
    seller = invoice_data.get("seller_details", {})
    buyer = invoice_data.get("buyer_details", {})

    seller_state_code = _safe_state_code(seller.get("state_code"))
    buyer_state_code = _safe_state_code(buyer.get("state_code"))
    is_intrastate = bool(seller_state_code and buyer_state_code and seller_state_code == buyer_state_code)

    prepared: list[dict[str, Any]] = []
    totals = {
        "taxable": Decimal("0.00"),
        "cgst": Decimal("0.00"),
        "sgst": Decimal("0.00"),
        "igst": Decimal("0.00"),
        "grand": Decimal("0.00"),
    }

    for index, item in enumerate(invoice_data.get("items", []), start=1):
        qty = _d(item.get("qty", 1))
        rate = _d(item.get("rate", 0))
        gst_rate = _d(item.get("gst_rate", 18))
        taxable = (qty * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if is_intrastate:
            cgst_rate = (gst_rate / Decimal("2")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            sgst_rate = cgst_rate
            igst_rate = Decimal("0.00")
        else:
            cgst_rate = Decimal("0.00")
            sgst_rate = Decimal("0.00")
            igst_rate = gst_rate

        cgst_amount = (taxable * cgst_rate / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        sgst_amount = (taxable * sgst_rate / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        igst_amount = (taxable * igst_rate / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        line_total = (taxable + cgst_amount + sgst_amount + igst_amount).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        totals["taxable"] += taxable
        totals["cgst"] += cgst_amount
        totals["sgst"] += sgst_amount
        totals["igst"] += igst_amount
        totals["grand"] += line_total

        prepared.append(
            {
                "sr_no": index,
                "name": item.get("name", ""),
                "uom": item.get("uom", "Nos."),
                "ac_code": item.get("ac_code", ""),
                "rate": rate,
                "taxable": taxable,
                "cgst_rate": cgst_rate,
                "cgst_amount": cgst_amount,
                "sgst_rate": sgst_rate,
                "sgst_amount": sgst_amount,
                "igst_rate": igst_rate,
                "igst_amount": igst_amount,
                "total": line_total,
            }
        )

    totals = {k: v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) for k, v in totals.items()}
    return prepared, totals, is_intrastate


def _setup_page(ws) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3, header=0.2, footer=0.2)
    ws.print_title_rows = "1:18"


def generate_gst_invoice_xlsx(invoice_data: dict[str, Any], output_dir: str | Path = ".") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Invoice"
    _setup_page(ws)

    for col, width in {
        "A": 6,
        "B": 30,
        "C": 8,
        "D": 10,
        "E": 11,
        "F": 13,
        "G": 8,
        "H": 13,
        "I": 8,
        "J": 13,
        "K": 8,
        "L": 13,
        "M": 14,
    }.items():
        ws.column_dimensions[col].width = width

    header_font = Font(name="Times New Roman", size=16, bold=True)
    title_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=10)
    body_bold = Font(name="Calibri", size=10, bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)

    seller = invoice_data.get("seller_details", {})
    buyer = invoice_data.get("buyer_details", {})
    bank = invoice_data.get("bank_details", {})

    company_name = str(seller.get("company_name") or "R. R. ENTERPRISES").upper()
    address_lines = seller.get("address_lines") or [
        "ADDRESS-C-102/A3) PHASE -7 INDL AREA MOHALI",
        "Mob. No. : 9888483817-9592410333",
        f"GSTIN : {seller.get('gstin', '')}",
        f"E-Mail : {seller.get('email', '')}",
    ]
    if isinstance(address_lines, str):
        address_lines = [address_lines]

    # Header section
    _merge_and_set(ws, 1, 1, 1, 13, company_name, font=header_font, align=centered)
    ws.row_dimensions[1].height = 28
    row_idx = 2
    for line in address_lines[:4]:
        _merge_and_set(ws, row_idx, row_idx, 1, 13, line, font=body_bold, align=centered)
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # Invoice information section (two-column)
    info_start = 6
    info_end = 11
    _set_border_range(ws, info_start, info_end, 1, 13)

    _set_cell(ws, info_start, 1, "Invoice No.", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start,
        info_start,
        2,
        6,
        str(invoice_data.get("invoice_no", "")),
        font=body_font,
        align=left_wrap,
    )

    _set_cell(ws, info_start + 1, 1, "Invoice Date", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start + 1,
        info_start + 1,
        2,
        6,
        _format_invoice_date(invoice_data.get("invoice_date")),
        font=body_font,
        align=left_wrap,
    )

    _set_cell(ws, info_start + 2, 1, "State", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start + 2,
        info_start + 2,
        2,
        3,
        str(seller.get("state", "")),
        font=body_font,
        align=left_wrap,
    )
    _set_cell(ws, info_start + 2, 4, "State Code", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start + 2,
        info_start + 2,
        5,
        6,
        _safe_state_code(seller.get("state_code", "")),
        font=body_font,
        align=left_wrap,
    )

    _merge_and_set(ws, info_start, info_start, 7, 13, "Details of Receiver", font=title_font, align=centered)
    _set_cell(ws, info_start + 1, 7, "Name", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws, info_start + 1, info_start + 1, 8, 13, str(buyer.get("name", "")), font=body_font, align=left_wrap
    )
    _set_cell(ws, info_start + 2, 7, "Address", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start + 2,
        info_start + 2,
        8,
        13,
        str(buyer.get("address", "")),
        font=body_font,
        align=left_wrap,
    )
    _set_cell(ws, info_start + 3, 7, "GSTIN", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start + 3,
        info_start + 3,
        8,
        13,
        str(buyer.get("gstin", "")),
        font=body_font,
        align=left_wrap,
    )
    _set_cell(ws, info_start + 4, 7, "State", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws, info_start + 4, info_start + 4, 8, 10, str(buyer.get("state", "")), font=body_font, align=left_wrap
    )
    _set_cell(ws, info_start + 4, 11, "State Code", font=body_bold, align=left_wrap)
    _merge_and_set(
        ws,
        info_start + 4,
        info_start + 4,
        12,
        13,
        _safe_state_code(buyer.get("state_code", "")),
        font=body_font,
        align=left_wrap,
    )

    items, totals, _ = _prepare_items(invoice_data)

    # Item table header section
    table_header_row_1 = 12
    table_header_row_2 = 13

    ws.row_dimensions[table_header_row_1].height = 22
    ws.row_dimensions[table_header_row_2].height = 22

    _merge_and_set(ws, table_header_row_1, table_header_row_2, 1, 1, "Sr. No", font=body_bold, align=centered)
    _merge_and_set(
        ws,
        table_header_row_1,
        table_header_row_2,
        2,
        2,
        "Name of Product/Service",
        font=body_bold,
        align=centered,
    )
    _merge_and_set(ws, table_header_row_1, table_header_row_2, 3, 3, "UOM", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_2, 4, 4, "AC CODE", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_2, 5, 5, "Rate", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_2, 6, 6, "Taxable Value", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_1, 7, 8, "CGST", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_1, 9, 10, "SGST", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_1, 11, 12, "IGST", font=body_bold, align=centered)
    _merge_and_set(ws, table_header_row_1, table_header_row_2, 13, 13, "Total (₹)", font=body_bold, align=centered)

    _set_cell(ws, table_header_row_2, 7, "Rate", font=body_bold, align=centered)
    _set_cell(ws, table_header_row_2, 8, "Amount", font=body_bold, align=centered)
    _set_cell(ws, table_header_row_2, 9, "Rate", font=body_bold, align=centered)
    _set_cell(ws, table_header_row_2, 10, "Amount", font=body_bold, align=centered)
    _set_cell(ws, table_header_row_2, 11, "Rate", font=body_bold, align=centered)
    _set_cell(ws, table_header_row_2, 12, "Amount", font=body_bold, align=centered)

    for c in range(1, 14):
        ws.cell(row=table_header_row_1, column=c).border = THIN_BORDER
        ws.cell(row=table_header_row_2, column=c).border = THIN_BORDER

    # Item rows
    item_start = 14
    for i, item in enumerate(items):
        row = item_start + i
        _set_cell(ws, row, 1, item["sr_no"], font=body_font, align=centered)
        _set_cell(ws, row, 2, item["name"], font=body_font, align=left_wrap)
        _set_cell(ws, row, 3, item["uom"], font=body_font, align=centered)
        _set_cell(ws, row, 4, item["ac_code"], font=body_font, align=centered)
        _set_cell(ws, row, 5, float(item["rate"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 6, float(item["taxable"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 7, float(item["cgst_rate"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 8, float(item["cgst_amount"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 9, float(item["sgst_rate"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 10, float(item["sgst_amount"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 11, float(item["igst_rate"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 12, float(item["igst_amount"]), font=body_font, align=right_wrap)
        _set_cell(ws, row, 13, float(item["total"]), font=body_font, align=right_wrap)

    if not items:
        _set_border_range(ws, item_start, item_start + 1, 1, 13)
        _merge_and_set(ws, item_start, item_start + 1, 2, 13, "", font=body_font, align=left_wrap)

    total_row = item_start + max(1, len(items))
    _merge_and_set(ws, total_row, total_row, 1, 5, "Total", font=body_bold, align=right_wrap)
    _set_cell(ws, total_row, 6, float(totals["taxable"]), font=body_bold, align=right_wrap)
    _set_cell(ws, total_row, 7, "", font=body_bold, align=centered)
    _set_cell(ws, total_row, 8, float(totals["cgst"]), font=body_bold, align=right_wrap)
    _set_cell(ws, total_row, 9, "", font=body_bold, align=centered)
    _set_cell(ws, total_row, 10, float(totals["sgst"]), font=body_bold, align=right_wrap)
    _set_cell(ws, total_row, 11, "", font=body_bold, align=centered)
    _set_cell(ws, total_row, 12, float(totals["igst"]), font=body_bold, align=right_wrap)
    _set_cell(ws, total_row, 13, float(totals["grand"]), font=body_bold, align=right_wrap)
    _set_border_range(ws, total_row, total_row, 1, 13)

    # Bottom section
    bottom_start = total_row + 1
    bottom_end = bottom_start + 6
    _set_border_range(ws, bottom_start, bottom_end, 1, 13)

    amount_words = amount_to_words_indian(totals["grand"])
    _merge_and_set(
        ws,
        bottom_start,
        bottom_start,
        1,
        7,
        f"Total Invoice Amount in Words : {amount_words}",
        font=body_bold,
        align=left_wrap,
    )

    bank_lines = [
        f"Bank Name : {bank.get('bank_name', '')}",
        f"Branch : {bank.get('branch', '')}",
        f"IFSC Code : {bank.get('ifsc_code', '')}",
        f"A/C No : {bank.get('account_no', '')}",
        f"Type of A/C : {bank.get('account_type', '')}",
    ]
    for offset, line in enumerate(bank_lines, start=1):
        _merge_and_set(ws, bottom_start + offset, bottom_start + offset, 1, 7, line, font=body_font, align=left_wrap)

    _merge_and_set(
        ws,
        bottom_end,
        bottom_end,
        1,
        7,
        "Terms & Conditions: Goods once sold will not be taken back. Please settle dues within 15 days.",
        font=body_font,
        align=left_wrap,
    )

    tax_rows = [
        ("Total Amount before Tax", totals["taxable"]),
        ("Add: CGST", totals["cgst"]),
        ("Add: SGST", totals["sgst"]),
        ("Add: IGST", totals["igst"]),
        ("Total Amount after Tax", totals["grand"]),
    ]

    for i, (label, value) in enumerate(tax_rows):
        row = bottom_start + i
        _merge_and_set(ws, row, row, 8, 12, label, font=body_bold if i in (0, 4) else body_font, align=left_wrap)
        _set_cell(
            ws,
            row,
            13,
            float(value),
            font=body_bold if i in (0, 4) else body_font,
            align=right_wrap,
        )

    # Signature section
    sign_start = bottom_end + 1
    sign_end = sign_start + 2
    _set_border_range(ws, sign_start, sign_end, 1, 13)

    _merge_and_set(ws, sign_start, sign_start, 8, 13, "For R. R. ENTERPRISES", font=body_bold, align=right_wrap)
    _merge_and_set(ws, sign_start + 1, sign_start + 1, 8, 13, "", font=body_font, align=right_wrap)
    _merge_and_set(ws, sign_end, sign_end, 8, 13, "Authorised Signatory", font=body_bold, align=right_wrap)

    # Global style tuning
    for row in ws.iter_rows(min_row=1, max_row=sign_end, min_col=1, max_col=13):
        for cell in row:
            if not cell.font:
                cell.font = body_font
            if not cell.alignment:
                cell.alignment = left_wrap

    for row in range(item_start, total_row + 1):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[bottom_start].height = 24
    ws.row_dimensions[bottom_end].height = 32

    invoice_no = str(invoice_data.get("invoice_no", "invoice")).replace("/", "-").replace("\\", "-")
    output_path = Path(output_dir) / f"invoice_{invoice_no}.xlsx"
    wb.save(output_path)
    return output_path
